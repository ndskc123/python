import openpyxl
import MySQLdb

book=openpyxl.load_workbook("./a.xlsx")
sheet=book['Sheet1']
list1=[]
list2=[]
max_row=sheet.max_row
max_column=sheet.max_column

for row in range(1,max_row+1):
    for column in range(1,max_column+1):
        date=sheet.cell(row,column).value
        list1.append(date)

for i in range(0,len(list1),8):
    b=list1[i:i+8]
    list2.append(b)

db=MySQLdb.connect(host='localhost',user='root',passwd='123456',port=3306,charset='utf8')
cursor=db.cursor()
cursor.execute("show databases;")
cursor.execute("use test;")
print(list2)
for i in range(0,9):
    cursor.execute("insert into xsb (xh,xm,xb,nl,bj,jg,sfzh,zcrq) values('%s','%s','%s','%s','%s','%s','%s','%s');"%(list2[i][0],list2[i][1],list2[i][2],list2[i][3],list2[i][4],list2[i][5],list2[i][6],list2[i][7]))
db.commit()
cursor.execute("select * from xsb;")
for date in cursor:
    print(date)

book.close()